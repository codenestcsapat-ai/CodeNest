#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
CodeNest outreach HTML generator.

Reads an Excel lead list and generates one personalized HTML email preview per valid lead
from a master HTML template.

Example:
python email-preview/generate_outreach_from_excel.py ^
  --excel email-preview/codenest_osszesitett_leadlista.xlsx ^
  --template email-preview/codenest-outreach-v2-positive-template.html ^
  --out-dir email-preview/generated ^
  --base-url https://codenest.hu/email-preview/generated ^
  --max 3
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import html
import re
import sys
from pathlib import Path
from urllib.parse import quote

from openpyxl import load_workbook


SHEET_NAME = "Első kör javaslat"
FROM_EMAIL = "info.codenest.hu@gmail.com"

FORBIDDEN_PHRASES = [
    "nem kész ajánlat",
    "nem sablonos",
    "rossz",
    "hiba",
    "hiányos",
    "problémás",
    "ráerőltet",
]


def clean_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def is_valid_email(value: str) -> bool:
    email = clean_text(value).lower()
    if not email:
        return False

    invalid_markers = [
        "not found",
        "nincs",
        "nem talál",
        "nem talal",
        "ismeretlen",
        "unknown",
        "obfuscated",
        "ellenőriz",
        "ellenoriz",
        "-",
    ]
    if any(marker in email for marker in invalid_markers):
        return False

    return re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email) is not None


def slugify(value: str, fallback: str = "lead") -> str:
    value = clean_text(value).lower()

    replacements = {
        "á": "a", "é": "e", "í": "i", "ó": "o", "ö": "o", "ő": "o",
        "ú": "u", "ü": "u", "ű": "u",
    }
    for source, target in replacements.items():
        value = value.replace(source, target)

    value = re.sub(r"[^a-z0-9]+", "-", value)
    value = re.sub(r"-+", "-", value).strip("-")

    return value or fallback


def short_company_name(company_name: str) -> str:
    name = clean_text(company_name)

    suffix_pattern = re.compile(
        r"\b(kft|kft\.|bt|bt\.|zrt|zrt\.|nyrt|nyrt\.|ev|e\.v\.|egyéni vállalkozó|korlátolt felelősségű társaság)\b",
        re.IGNORECASE,
    )
    name = suffix_pattern.sub("", name)
    name = re.sub(r"\s+", " ", name).strip(" ,.-–—")

    if not name:
        return clean_text(company_name)

    # Keep it reasonably short for subject lines and filenames.
    return name[:60].strip(" ,.-–—")


def token_for_row(company_name: str, email: str) -> str:
    raw = f"{company_name}|{email}".encode("utf-8")
    return hashlib.sha256(raw).hexdigest()[:10]


def category_context(company_type: str) -> str:
    text = clean_text(company_type).lower()

    if any(word in text for word in ["gyárt", "ipar", "b2b", "műanyag", "logiszt", "technológ", "termel"]):
        return "gyártócéges / B2B működés"

    if any(word in text for word in ["turizmus", "szállás", "hotel", "vendég", "étterem", "kávézó", "panzió"]):
        return "turisztikai vagy vendéglátóipari működés"

    if any(word in text for word in ["önkormányzat", "intézmény", "iskola", "egyesület", "alapítvány", "szervezet"]):
        return "intézményi vagy közösségi működés"

    if any(word in text for word in ["szolgált", "keresked", "bolt", "üzlet", "szalon", "rendelő"]):
        return "helyi szolgáltatói működés"

    return "céges működés"


def local_context(city_region: str) -> str:
    value = clean_text(city_region)
    if value:
        return f"{value} környéki jelenlét"
    return "helyi jelenlét"


def personal_note(row: dict) -> str:
    note = clean_text(row.get("Személyes nyitás ötlet", ""))
    angle = clean_text(row.get("Javasolt CodeNest angle", ""))

    # Prefer the explicit personal note if it exists.
    if note:
        return note

    if angle:
        return f"Az első benyomás alapján a legfontosabb irány a következő lehet: {angle}"

    return "Egy átláthatóbb online jelenlét több ponton is segítheti a cég bemutatását, a kapcsolatfelvételt és a későbbi fejlesztési lehetőségek felépítését."


def make_mailto(company_short: str) -> str:
    subject = f"{company_short} online jelenléte – egy rövid ötlet"
    body = (
        "Kedves Bors!\n\n"
        "Köszönjük a megkeresést. Nyitottak vagyunk egy rövid egyeztetésre.\n\n"
        "Nekünk ezek az időpontok lennének megfelelőek:\n"
        "1. \n"
        "2. \n"
        "3. \n\n"
        "Üdvözlettel,\n"
    )

    return f"mailto:{FROM_EMAIL}?subject={quote(subject)}&body={quote(body)}"


def normalize_base_url(base_url: str) -> str:
    return clean_text(base_url).rstrip("/")


def replace_placeholders(template: str, replacements: dict[str, str]) -> str:
    rendered = template
    for key, value in replacements.items():
        rendered = rendered.replace("{{" + key + "}}", html.escape(value, quote=False))
    return rendered


def remaining_placeholders(rendered: str) -> list[str]:
    return sorted(set(re.findall(r"\{\{[A-Z0-9_]+\}\}", rendered)))


def contains_forbidden_text(rendered: str) -> list[str]:
    lower = rendered.lower()
    return [phrase for phrase in FORBIDDEN_PHRASES if phrase in lower]


def load_rows(excel_path: Path) -> list[dict]:
    wb = load_workbook(excel_path, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise ValueError(f"Missing sheet '{SHEET_NAME}'. Available sheets: {available}")

    ws = wb[SHEET_NAME]
    headers = [clean_text(cell.value) for cell in ws[1]]

    required_headers = ["Név", "Email"]
    missing = [header for header in required_headers if header not in headers]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        item = {}
        for header, value in zip(headers, row):
            if header:
                item[header] = value
        rows.append(item)

    return rows


def generate(args: argparse.Namespace) -> int:
    excel_path = Path(args.excel)
    template_path = Path(args.template)
    out_dir = Path(args.out_dir)
    base_url = normalize_base_url(args.base_url)

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    if not template_path.exists():
        raise FileNotFoundError(f"Template file not found: {template_path}")

    out_dir.mkdir(parents=True, exist_ok=True)

    template = template_path.read_text(encoding="utf-8")
    rows = load_rows(excel_path)

    manifest_rows = []
    skipped_rows = []
    generated_count = 0

    for row in rows:
        if args.max is not None and generated_count >= args.max:
            break

        company_name = clean_text(row.get("Név", ""))
        email = clean_text(row.get("Email", ""))
        status = clean_text(row.get("Státusz", ""))

        if not company_name:
            skipped_rows.append({
                "company_name": company_name,
                "email": email,
                "reason": "missing company name",
            })
            continue

        if not is_valid_email(email):
            skipped_rows.append({
                "company_name": company_name,
                "email": email,
                "reason": "missing or invalid email",
            })
            continue

        company_short = short_company_name(company_name)
        token = token_for_row(company_name, email)
        filename = f"{slugify(company_short)}-{token}.html"
        preview_url = f"{base_url}/{filename}"
        subject = f"{company_short} online jelenléte – egy rövid ötlet"

        replacements = {
            "COMPANY_NAME": company_name,
            "COMPANY_SHORT": company_short,
            "PERSONAL_NOTE": personal_note(row),
            "INDUSTRY_CONTEXT": category_context(clean_text(row.get("Típus", ""))),
            "LOCAL_CONTEXT": local_context(clean_text(row.get("Város / régió", ""))),
            "PREHEADER": f"Egy rövid ötlet {company_short} online jelenlétéhez.",
            "PREVIEW_URL": preview_url,
            "REPLY_MAILTO": make_mailto(company_short),
        }

        rendered = replace_placeholders(template, replacements)

        placeholders = remaining_placeholders(rendered)
        forbidden = contains_forbidden_text(rendered)

        if placeholders:
            skipped_rows.append({
                "company_name": company_name,
                "email": email,
                "reason": f"unresolved placeholders: {', '.join(placeholders)}",
            })
            continue

        if forbidden:
            skipped_rows.append({
                "company_name": company_name,
                "email": email,
                "reason": f"forbidden phrase found: {', '.join(forbidden)}",
            })
            continue

        # The generated HTML must not expose the recipient email.
        if email and email in rendered:
            skipped_rows.append({
                "company_name": company_name,
                "email": email,
                "reason": "recipient email leaked into HTML",
            })
            continue

        (out_dir / filename).write_text(rendered, encoding="utf-8")

        manifest_rows.append({
            "company_name": company_name,
            "company_short": company_short,
            "email": email,
            "subject": subject,
            "filename": filename,
            "preview_url": preview_url,
            "status": status,
        })
        generated_count += 1

    manifest_path = out_dir / "manifest_private.csv"
    skipped_path = out_dir / "skipped_private.csv"

    with manifest_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["company_name", "company_short", "email", "subject", "filename", "preview_url", "status"],
        )
        writer.writeheader()
        writer.writerows(manifest_rows)

    with skipped_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=["company_name", "email", "reason"])
        writer.writeheader()
        writer.writerows(skipped_rows)

    print(f"Generated HTML files: {generated_count}")
    print(f"Skipped rows: {len(skipped_rows)}")
    print(f"Output folder: {out_dir}")
    print(f"Manifest: {manifest_path}")
    print(f"Skipped: {skipped_path}")

    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Generate personalized CodeNest outreach HTML emails from Excel.")
    parser.add_argument("--excel", required=True, help="Path to the Excel lead list.")
    parser.add_argument("--template", required=True, help="Path to the master HTML template.")
    parser.add_argument("--out-dir", required=True, help="Output folder for generated HTML and CSV files.")
    parser.add_argument("--base-url", required=True, help="Public base URL for generated preview links.")
    parser.add_argument("--max", type=int, default=None, help="Optional maximum number of generated valid emails.")
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    return generate(args)


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1)